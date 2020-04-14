import openpyxl
from translate import Translator

#open workbook
wb = openpyxl.load_workbook('Python Translate.xlsx')

#label individual sheets
ws1 = wb['Sheet1']
ws2 = wb['Sheet2']

#start at cell A1
row_loc = 1
column_loc = 1
#record cell values

while column_loc <= 10:
    
    while row_loc <= 10:

        if ws1.cell(row = row_loc, column = column_loc).value != None:
            text = ws1.cell(row = row_loc, column = column_loc).value


            t = Translator(from_lang='English', to_lang='Spanish')
            translation = t.translate(text)

            ws2.cell(row = row_loc, column = column_loc).value = translation
            wb.save('Python Translate.xlsx')
        row_loc += 1
        
    column_loc += 1

print("done")


